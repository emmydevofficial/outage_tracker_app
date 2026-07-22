"""Read Benin Region extra sheets and remaining data."""
import openpyxl
fpath = r"C:\Users\HP\OneDrive\Desktop\TCN Files\Benin Region ISO Parameter.xlsx"
wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
# Read all sheets that haven't been fully shown
for sname in wb.sheetnames:
    ws = wb[sname]
    if sname == "Parameters 6":
        # Read rows beyond 80 for this sheet
        for i, row in enumerate(ws.iter_rows(max_col=5, values_only=True), 1):
            if i <= 80:
                continue
            cells = [str(c).strip() if c else "" for c in row]
            if all(x == "" for x in cells):
                continue
            line = " | ".join(cells)
            print(f"  {sname} R{i:3d}: {line}")
wb.close()
