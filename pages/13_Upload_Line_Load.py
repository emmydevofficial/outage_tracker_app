"""
### FILE: pages/13_Upload_Line_Load.py
Upload the (330/132kV LINE) Load Management Tracking sheet for a region and
upsert it into the line_load table. Region is read from the sheet itself
(cell A1), so the same page handles any region's file.

Sheet layout expected:
  A1            merged region name (e.g. "OSOGBO")
  Row 2, L:AI   hourly time headers ("01:00" .. "24:00")
  Column B      ACC (vertically merged across rows sharing the same ACC)
  Column C      Transmission Interface (vertically merged)
  Column D      Line Voltage (vertically merged)
  Column E      Line Nomenclature (one row per line)
  Column K      Disco
  Row 3..end, L:AI   hourly load readings for that line
"""

import streamlit as st
from utils.auth import login, is_super_admin, current_region

login()

import pandas as pd
from openpyxl import load_workbook
from utils.db import upsert_line_load, read_line_load
from utils.load_upload import build_merge_fill_map, normalize_time_header, classify_cell
from utils.regions import normalize_region
from utils.activity_log import log_activity
from utils.file_storage import save_uploaded_file

st.set_page_config(page_title="Upload Line Load", layout="wide")

st.title("📥 Upload 330/132kV Line Load Tracking")
st.markdown(
    "Upload the hourly (330/132kV LINE) Load Management Tracking sheet for a single region/date. "
    "The region is read from the sheet itself, so this works for any region's file."
)

ACC_COL = 2                      # column B
TRANSMISSION_INTERFACE_COL = 3   # column C
LINE_VOLTAGE_COL = 4             # column D
LINE_NOMENCLATURE_COL = 5        # column E
DISCO_COL = 11                   # column K
FIRST_HOUR_COL = 12              # column L
LAST_HOUR_COL = 35               # column AI
DATA_START_ROW = 3

reading_date = st.date_input("Reading Date")
uploaded = st.file_uploader("Choose Line Load Excel file", type=["xlsx", "xlsm"])

if uploaded is not None:
    try:
        wb = load_workbook(uploaded, data_only=True)
    except Exception as e:
        st.error(f"Could not read the Excel file: {e}")
        st.stop()

    ws = wb.worksheets[0]

    region = ws.cell(row=1, column=1).value
    if not region or not str(region).strip():
        st.error("Could not find a region name in cell A1 of the uploaded sheet.")
        st.stop()
    region = normalize_region(region)

    if not is_super_admin() and region.strip().upper() != str(current_region()).strip().upper():
        st.error(f"This file is for **{region}**, but you only have upload access to **{current_region()}**.")
        st.stop()

    hour_headers = [
        normalize_time_header(ws.cell(row=2, column=col).value)
        for col in range(FIRST_HOUR_COL, LAST_HOUR_COL + 1)
    ]
    if any(not h for h in hour_headers):
        st.error("One or more hourly column headers (row 2, columns L:AI) are blank or unreadable.")
        st.stop()

    area_by_row = build_merge_fill_map(ws, ACC_COL, ws.max_row)
    interface_by_row = build_merge_fill_map(ws, TRANSMISSION_INTERFACE_COL, ws.max_row)
    voltage_by_row = build_merge_fill_map(ws, LINE_VOLTAGE_COL, ws.max_row)

    rows = []
    kinds = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        line_nomenclature = ws.cell(row=r, column=LINE_NOMENCLATURE_COL).value
        if line_nomenclature is None or str(line_nomenclature).strip() == "":
            continue
        line_nomenclature = str(line_nomenclature).strip()

        area = area_by_row.get(r)
        area = str(area).strip() if area is not None else None
        transmission_interface = interface_by_row.get(r)
        transmission_interface = str(transmission_interface).strip() if transmission_interface is not None else None
        line_voltage = voltage_by_row.get(r)
        line_voltage = str(line_voltage).strip() if line_voltage is not None else None
        disco = ws.cell(row=r, column=DISCO_COL).value
        disco = str(disco).strip() if disco is not None else None

        for i, col in enumerate(range(FIRST_HOUR_COL, LAST_HOUR_COL + 1)):
            raw = ws.cell(row=r, column=col).value
            load_mw, cause, kind = classify_cell(raw)
            rows.append({
                "reading_date": reading_date,
                "reading_time": hour_headers[i],
                "region": region,
                "area": area,
                "transmission_interface": transmission_interface,
                "disco": disco,
                "line_voltage": line_voltage,
                "line_nomenclature": line_nomenclature,
                "load_mw": load_mw,
                "cause": cause,
            })
            kinds.append(kind)

    if not rows:
        st.warning("No line rows found in the uploaded sheet (column E was empty throughout).")
        st.stop()

    df = pd.DataFrame(rows)
    kind_counts = pd.Series(kinds).value_counts()

    st.subheader("Parse Summary")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Region", region)
    c2.metric("Total Readings", f"{len(df):,}")
    c3.metric("Numeric Readings", f"{kind_counts.get('numeric', 0):,}")
    c4.metric("Flagged (fault/format)", f"{kind_counts.get('fault', 0) + kind_counts.get('wrong_format', 0):,}")

    flagged = df[df["cause"].notna()]
    if not flagged.empty:
        with st.expander(f"View {len(flagged):,} flagged reading(s) (fault codes + wrong data format)"):
            st.dataframe(
                flagged[["area", "transmission_interface", "line_voltage", "line_nomenclature", "disco", "reading_time", "cause"]],
                use_container_width=True,
                height=300,
            )

    st.subheader("Preview (first 50 rows)")
    st.dataframe(df.head(50), use_container_width=True)
    st.caption(
        "Re-uploading the same reading_date/reading_time/transmission_interface/line_nomenclature "
        "combination will update the existing row instead of creating a duplicate."
    )

    if st.button("Upload to database", type="primary"):
        try:
            upsert_line_load(df)
            read_line_load.clear()
            st.success(f"{len(df):,} line load reading(s) for {region} on {reading_date} uploaded successfully.")
            save_uploaded_file(uploaded, "Upload Line Load", region)
            log_activity("upload_line_load", f"Uploaded '{uploaded.name}' — {len(df):,} reading(s) for {region} on {reading_date}")
        except Exception as e:
            st.error(f"Error inserting records: {e}")
