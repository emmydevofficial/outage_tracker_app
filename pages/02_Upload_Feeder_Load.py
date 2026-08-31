"""
### FILE: pages/02_Upload_Feeder_Load.py
Upload the DISCO (33kV Feeder) Load Management Tracking sheet for a region
and upsert it into the feeder_33kv_load table. Region is read from the sheet
itself (cell A1), so the same page handles any region's file.

Sheet layout expected:
  A1            merged region name (e.g. "OSOGBO")
  Row 2, M:AJ   hourly time headers ("01:00" .. "24:00")
  Column B      ACC (vertically merged across rows sharing the same ACC)
  Column C      Station / Transmission Interface (vertically merged)
  Column F      33kV feeder name (one row per feeder)
  Column L      Customer / Disco
  Row 3..end, M:AJ   hourly load readings for that feeder
"""

import streamlit as st
from utils.auth import login, is_super_admin, current_region

login()

import pandas as pd
from openpyxl import load_workbook
from utils.db import upsert_feeder_load, read_feeder_load
from utils.load_upload import build_merge_fill_map, normalize_time_header, classify_cell
from utils.regions import normalize_region
from utils.activity_log import log_activity
from utils.file_storage import save_uploaded_file
from utils.branding import inject_css, page_header, kpi_card, kpi_grid, one_indexed

st.set_page_config(page_title="Upload Feeder Load", page_icon="⚡", layout="wide")
inject_css()
page_header("Upload 33kV Feeder Load Tracking", "33kV Feeder Network · Data Ingestion")
st.markdown(
    "Upload the hourly DISCO (33kV Feeder) Load Management Tracking sheet for a single region/date. "
    "The region is read from the sheet itself, so this works for any region's file."
)

ACC_COL = 2         # column B
STATION_COL = 3     # column C
FEEDER_COL = 6      # column F
CUSTOMER_COL = 12   # column L
FIRST_HOUR_COL = 13  # column M
LAST_HOUR_COL = 36   # column AJ
DATA_START_ROW = 3

reading_date = st.date_input("Reading Date")
uploaded = st.file_uploader("Choose Feeder Load Excel file", type=["xlsx", "xlsm"])

if uploaded is not None:
    try:
        wb = load_workbook(uploaded, data_only=True)
    except Exception as e:
        st.error(f"Could not read the Excel file: {e}")
        st.stop()

    ws = wb.worksheets[0]

    region = ws.cell(row=1, column=1).value
    if not region or not str(region).strip():
        st.error("Could not find a region name in cell A1 of the uploaded sheet.")
        st.stop()
    region = normalize_region(region)

    if not is_super_admin() and region.strip().upper() != str(current_region()).strip().upper():
        st.error(f"This file is for **{region}**, but you only have upload access to **{current_region()}**.")
        st.stop()

    hour_headers = [
        normalize_time_header(ws.cell(row=2, column=col).value)
        for col in range(FIRST_HOUR_COL, LAST_HOUR_COL + 1)
    ]
    if any(not h for h in hour_headers):
        st.error("One or more hourly column headers (row 2, columns M:AJ) are blank or unreadable.")
        st.stop()

    area_by_row = build_merge_fill_map(ws, ACC_COL, ws.max_row)
    station_by_row = build_merge_fill_map(ws, STATION_COL, ws.max_row)

    rows = []
    kinds = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        feeder = ws.cell(row=r, column=FEEDER_COL).value
        if feeder is None or str(feeder).strip() == "":
            continue
        feeder = str(feeder).strip()

        area = area_by_row.get(r)
        area = str(area).strip() if area is not None else None
        station = station_by_row.get(r)
        station = str(station).strip() if station is not None else None
        customer = ws.cell(row=r, column=CUSTOMER_COL).value
        customer = str(customer).strip() if customer is not None else None

        for i, col in enumerate(range(FIRST_HOUR_COL, LAST_HOUR_COL + 1)):
            raw = ws.cell(row=r, column=col).value
            load_mw, cause, kind = classify_cell(raw)
            rows.append({
                "reading_date": reading_date,
                "reading_time": hour_headers[i],
                "region": region,
                "area": area,
                "feeder": feeder,
                "customer": customer,
                "station": station,
                "load_mw": load_mw,
                "cause": cause,
            })
            kinds.append(kind)

    if not rows:
        st.warning("No feeder rows found in the uploaded sheet (column F was empty throughout).")
        st.stop()

    df = pd.DataFrame(rows)
    kind_counts = pd.Series(kinds).value_counts()

    st.subheader("Parse Summary")
    kpi_grid([
        kpi_card("Region", region, "", "building", "#1e3a7a"),
        kpi_card("Total Readings", f"{len(df):,}", "", "chart", "#1F6C9F"),
        kpi_card("Numeric Readings", f"{kind_counts.get('numeric', 0):,}", "", "pulse", "#346538"),
        kpi_card("Flagged", f"{kind_counts.get('fault', 0) + kind_counts.get('wrong_format', 0):,}", "", "alert", "#c81e28"),
    ])

    flagged = df[df["cause"].notna()]
    if not flagged.empty:
        with st.expander(f"View {len(flagged):,} flagged reading(s) (fault codes + wrong data format)"):
            st.dataframe(
                one_indexed(flagged[["area", "station", "feeder", "customer", "reading_time", "cause"]]),
                use_container_width=True,
                height=300,
            )

    st.subheader("Preview (first 50 rows)")
    st.dataframe(one_indexed(df.head(50)), use_container_width=True)
    st.caption(
        "Re-uploading the same reading_date/reading_time/station/feeder combination "
        "will update the existing row instead of creating a duplicate."
    )

    if st.button("Upload to database", type="primary"):
        try:
            upsert_feeder_load(df)
            read_feeder_load.clear()
            st.success(f"{len(df):,} feeder load reading(s) for {region} on {reading_date} uploaded successfully.")
            save_uploaded_file(uploaded, "Upload Feeder Load", region)
            log_activity("upload_feeder_load", f"Uploaded '{uploaded.name}' — {len(df):,} reading(s) for {region} on {reading_date}")
        except Exception as e:
            st.error(f"Error inserting records: {e}")
