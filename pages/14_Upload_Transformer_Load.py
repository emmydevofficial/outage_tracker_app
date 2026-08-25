"""
### FILE: pages/14_Upload_Transformer_Load.py
Upload the TRANSFORMER HOURLY LOAD (MW) sheet for a region and upsert it into
the transformer_load table. Region is read from the sheet itself (cell A1),
so the same page handles any region's file.

Unlike the feeder/line load sheets, this one is transposed: transformers run
across the columns and hours run down the rows.

Sheet layout expected:
  A1                merged region name (e.g. "OSOGBO")
  Row 2, B:last      ACC (horizontally merged across each ACC's transformer columns)
  Row 3, B:last      Station (horizontally merged across each station's transformer columns)
  Row 5, B:last      Transformer Nomenclature (one column per transformer)
  Column A, 6:29     hour labels (e.g. "0100" .. "2400")
  B6:last,29         hourly load readings (rows = hours, columns = transformers)
"""

import streamlit as st
from utils.auth import login, is_super_admin, current_region

login()

import pandas as pd
from openpyxl import load_workbook
from utils.db import upsert_transformer_load, read_transformer_load
from utils.load_upload import build_merge_fill_map_row, normalize_hour_code, classify_cell
from utils.regions import normalize_region
from utils.activity_log import log_activity
from utils.file_storage import save_uploaded_file
from utils.branding import inject_css, page_header, kpi_card, kpi_grid

st.set_page_config(page_title="Upload Transformer Load", page_icon="⚡", layout="wide")
inject_css()
page_header("Upload Transformer Hourly Load", "33kV Feeder Network · Data Ingestion")
st.markdown(
    "Upload the TRANSFORMER HOURLY LOAD (MW) sheet for a single region/date. "
    "The region is read from the sheet itself, so this works for any region's file."
)

AREA_ROW = 2
STATION_ROW = 3
TRANSFORMER_ROW = 5
TIME_COL = 1        # column A
DATA_START_COL = 2  # column B
DATA_START_ROW = 6
DATA_END_ROW = 29    # 24 hourly rows: A6:A29

reading_date = st.date_input("Reading Date")
uploaded = st.file_uploader("Choose Transformer Load Excel file", type=["xlsx", "xlsm"])

if uploaded is not None:
    try:
        wb = load_workbook(uploaded, data_only=True)
    except Exception as e:
        st.error(f"Could not read the Excel file: {e}")
        st.stop()

    ws = wb.worksheets[0]

    region = ws.cell(row=1, column=1).value
    if not region or not str(region).strip():
        st.error("Could not find a region name in cell A1 of the uploaded sheet.")
        st.stop()
    region = normalize_region(region)

    if not is_super_admin() and region.strip().upper() != str(current_region()).strip().upper():
        st.error(f"This file is for **{region}**, but you only have upload access to **{current_region()}**.")
        st.stop()

    if ws.max_row < DATA_END_ROW:
        st.error(f"Expected hourly rows through row {DATA_END_ROW} (A6:A{DATA_END_ROW}); sheet only has {ws.max_row} rows.")
        st.stop()

    reading_times = [
        normalize_hour_code(ws.cell(row=r, column=TIME_COL).value)
        for r in range(DATA_START_ROW, DATA_END_ROW + 1)
    ]
    if any(not t for t in reading_times):
        st.error(f"One or more hour labels (column A, rows {DATA_START_ROW}:{DATA_END_ROW}) are blank or unreadable.")
        st.stop()

    # last column with a transformer nomenclature in row 5 (don't trust
    # ws.max_column blindly -- trailing formatted-but-empty columns happen)
    last_col = DATA_START_COL - 1
    for c in range(DATA_START_COL, ws.max_column + 1):
        v = ws.cell(row=TRANSFORMER_ROW, column=c).value
        if v is not None and str(v).strip() != "":
            last_col = c

    if last_col < DATA_START_COL:
        st.error(f"No transformer nomenclature found in row {TRANSFORMER_ROW} (expected starting at column B).")
        st.stop()

    area_by_col = build_merge_fill_map_row(ws, AREA_ROW, last_col)
    station_by_col = build_merge_fill_map_row(ws, STATION_ROW, last_col)

    rows = []
    kinds = []
    for c in range(DATA_START_COL, last_col + 1):
        transformer_nomenclature = ws.cell(row=TRANSFORMER_ROW, column=c).value
        if transformer_nomenclature is None or str(transformer_nomenclature).strip() == "":
            continue
        transformer_nomenclature = str(transformer_nomenclature).strip()

        area = area_by_col.get(c)
        area = str(area).strip() if area is not None else None
        station = station_by_col.get(c)
        station = str(station).strip() if station is not None else None

        for i, r in enumerate(range(DATA_START_ROW, DATA_END_ROW + 1)):
            raw = ws.cell(row=r, column=c).value
            load_mw, cause, kind = classify_cell(raw)
            rows.append({
                "reading_date": reading_date,
                "reading_time": reading_times[i],
                "region": region,
                "area": area,
                "station": station,
                "transformer_nomenclature": transformer_nomenclature,
                "load_mw": load_mw,
                "cause": cause,
            })
            kinds.append(kind)

    if not rows:
        st.warning("No transformer columns found in the uploaded sheet (row 5 was empty throughout).")
        st.stop()

    df = pd.DataFrame(rows)
    kind_counts = pd.Series(kinds).value_counts()

    st.subheader("Parse Summary")
    kpi_grid([
        kpi_card("Region", region, "", "building", "#1e3a7a"),
        kpi_card("Total Readings", f"{len(df):,}", "", "chart", "#1F6C9F"),
        kpi_card("Numeric Readings", f"{kind_counts.get('numeric', 0):,}", "", "pulse", "#346538"),
        kpi_card("Flagged", f"{kind_counts.get('fault', 0) + kind_counts.get('wrong_format', 0):,}", "", "alert", "#c81e28"),
    ])

    flagged = df[df["cause"].notna()]
    if not flagged.empty:
        with st.expander(f"View {len(flagged):,} flagged reading(s) (fault codes + wrong data format)"):
            st.dataframe(
                flagged[["area", "station", "transformer_nomenclature", "reading_time", "cause"]].reset_index(drop=True),
                use_container_width=True,
                height=300,
            )

    st.subheader("Preview (first 50 rows)")
    st.dataframe(df.head(50).reset_index(drop=True), use_container_width=True)
    st.caption(
        "Re-uploading the same reading_date/reading_time/station/transformer_nomenclature "
        "combination will update the existing row instead of creating a duplicate."
    )

    if st.button("Upload to database", type="primary"):
        try:
            upsert_transformer_load(df)
            read_transformer_load.clear()
            st.success(f"{len(df):,} transformer load reading(s) for {region} on {reading_date} uploaded successfully.")
            save_uploaded_file(uploaded, "Upload Transformer Load", region)
            log_activity("upload_transformer_load", f"Uploaded '{uploaded.name}' — {len(df):,} reading(s) for {region} on {reading_date}")
        except Exception as e:
            st.error(f"Error inserting records: {e}")
