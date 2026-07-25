"""
Extract transmission equipment names from TCN regional ISO parameter Excel files.
READ-ONLY - does not modify any files.
"""
import openpyxl
import os
import sys

FOLDER = r"C:\Users\HP\OneDrive\Desktop\TCN Files"

FILES = [
    "ABUJA ISO PARAMETERS.xlsx",
    "BAUCH REGION ISO PARAMETER.xlsx",
    "Benin Region ISO Parameter.xlsx",
    "ENUGU REGION SYSTEM ISO PARAMETERS (1).xlsx",
    "kaduna iso region Parameters.xlsx",
    "LAGOS REGION ISO Parameter.xlsx",
    "SHIRORO ISO parameter.xlsx",
    "PH REGION OPERATING PARAMETERS ON  TUESDAY 18TH MARCH 2025.xlsx",
    "OSOSOGBO REGION ISO PARAMETERS.xlsx",
    "KANO REGION ISO PARAMETERS.xlsx",
]

MAX_ROWS = 80  # rows per sheet
MAX_COLS = 5   # columns to show

for fname in FILES:
    fpath = os.path.join(FOLDER, fname)
    print("=" * 100)
    print(f"FILE: {fname}")
    print("=" * 100)
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n  --- Sheet: {sname} ---")
            row_count = 0
            for row in ws.iter_rows(max_col=MAX_COLS, values_only=True):
                if row_count >= MAX_ROWS:
                    print(f"    ... (truncated at {MAX_ROWS} rows)")
                    break
                # Format cells for display
                cells = []
                for c in row:
                    if c is None:
                        cells.append("")
                    else:
                        cells.append(str(c).strip())
                # Skip completely empty rows
                if all(x == "" for x in cells):
                    row_count += 1
                    continue
                line = " | ".join(cells)
                print(f"    R{row_count+1:3d}: {line}")
                row_count += 1
            if row_count == 0:
                print("    (empty sheet)")
        wb.close()
    except Exception as e:
        print(f"  ERROR: {e}")
    print()
